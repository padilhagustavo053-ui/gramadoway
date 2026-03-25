"""
Extrai produtos da planilha Gramadoway (Excel).
Ordem: variável de ambiente → pasta data/ → Área de trabalho (uso local).
"""
import io
import os
import re
from pathlib import Path
from typing import Any

from config_paths import data_root


def _pontuacao_nome_planilha(nome_arquivo: str) -> int:
    """
    Prioriza nomes típicos da planilha Gramadoway (evita pegar outro .xlsx qualquer).
    """
    n = nome_arquivo.lower()
    n = n.replace("í", "i").replace("ç", "c")
    score = 0
    if "gramadoway" in n:
        score += 12
    if "gramado" in n:
        score += 10
    if "planilha" in n:
        score += 4
    if "preco" in n or "preço" in nome_arquivo.lower():
        score += 5
    if "chocolate" in n or "bombom" in n or "barra" in n:
        score += 2
    return score


def _melhor_xlsx_em_pasta(pasta: Path) -> Path | None:
    """Entre vários .xlsx, escolhe o mais provável (nome + mais recente)."""
    todos = list(pasta.glob("*.xlsx"))
    if not todos:
        return None
    scored = [
        (_pontuacao_nome_planilha(f.name), f.stat().st_mtime, f)
        for f in todos
        if _pontuacao_nome_planilha(f.name) > 0
    ]
    if scored:
        scored.sort(key=lambda t: (-t[0], -t[1]))
        return scored[0][2]
    if len(todos) == 1:
        return todos[0]
    return None


def _iter_desktop_dirs() -> list[Path]:
    """Desktop local e Desktop do OneDrive (Windows comum)."""
    h = Path.home()
    out: list[Path] = []
    for rel in ("Desktop", "OneDrive/Desktop", "OneDrive/Área de Trabalho"):
        p = h / rel
        if p.is_dir():
            out.append(p)
    # Evita duplicatas se os dois apontarem ao mesmo sítio
    seen: set[Path] = set()
    uniq: list[Path] = []
    for p in out:
        try:
            r = p.resolve()
        except OSError:
            r = p
        if r not in seen:
            seen.add(r)
            uniq.append(p)
    return uniq


def _caminho_planilha() -> Path:
    """Localiza planilha: GRAMADOWAY_PLANILHA, depois data/, depois Desktop."""
    env = os.environ.get("GRAMADOWAY_PLANILHA", "").strip()
    if env:
        p = Path(env).expanduser()
        if p.is_file():
            return p

    root = data_root()
    for name in (
        "planilha.xlsx",
        "Planilha preços Gramadoway (1).xlsx",
        "Planilha precos Gramadoway (1).xlsx",
    ):
        cand = root / name
        if cand.is_file():
            return cand

    melhor_data = _melhor_xlsx_em_pasta(root)
    if melhor_data is not None:
        return melhor_data

    desktop_hint = Path.home() / "Desktop"
    for desktop in _iter_desktop_dirs():
        melhor = _melhor_xlsx_em_pasta(desktop)
        if melhor is not None:
            return melhor

    raise FileNotFoundError(
        f"Planilha .xlsx não encontrada. Opções:\n"
        f"• Coloque o arquivo em: {root} (ex.: planilha.xlsx)\n"
        f"• Ou defina GRAMADOWAY_PLANILHA com o caminho completo\n"
        f"• Ou use o upload na tela do app (nuvem)\n"
        f"• (PC local) Área de trabalho com nome tipo *Gramado* / *planilha* / *preços*\n"
        f"  Pastas verificadas: {', '.join(str(p) for p in _iter_desktop_dirs()) or desktop_hint}"
    )


def _parse_valor(val: Any) -> float | None:
    """Converte valor para float. Aceita número, 'R$ 99,00' ou 'R$ 99,00 Kg'."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val) if val else None
    s = str(val).strip()
    s = re.sub(r"R\$\s*", "", s, flags=re.I)
    s = s.replace(".", "").replace(",", ".")
    m = re.search(r"[\d.]+", s)
    if m:
        try:
            return float(m.group(0))
        except ValueError:
            pass
    return None


def _extrair_codigo(produto: str) -> tuple[str, str]:
    """Extrai código do início. Ex: '480- Avião 35g' -> ('480', 'Avião 35g')."""
    produto = str(produto).strip()
    m = re.match(r"^(\d+(?:\.\d+)?)\s*[-.]?\s*(.*)$", produto)
    if m:
        return m.group(1).strip(), (m.group(2).strip() or produto)
    return "", produto


def _linhas_dados(ws, padding: int = 8000, cap: int = 30000) -> range:
    """
    Intervalo de linhas a percorrer em cada aba.
    Amplia bem além de max_row (planilhas antigas / export às vezes reportam poucas linhas).
    Percorre colunas típicas e estende o fim até a última célula preenchida + margem.
    """
    mr = max(ws.max_row or 0, 3)
    scan_end = min(mr + padding, cap)
    last = mr
    for r in range(3, scan_end + 1):
        for c in (1, 2, 3, 6, 7, 8):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                last = r
                break
    end = min(last + 30, cap)
    return range(3, end + 1)


# --- ABA 1: Personalizados ---
# L1: título | L2: PRODUTO, VALOR, QTIDE, TOTAL | L3+: dados
# Col A=produto (pode ter código), Col B=valor (UN)
def extrair_personalizados(ws) -> list[dict]:
    itens = []
    for r in _linhas_dados(ws):
        prod = ws.cell(r, 1).value
        val = ws.cell(r, 2).value
        if not prod or str(prod).strip() in ("PRODUTO", "TOTAL", ""):
            continue
        preco = _parse_valor(val)
        if preco is None:
            preco = 0.0
        if preco < 0:
            continue
        cod, nome = _extrair_codigo(prod)
        itens.append({
            "categoria": "Personalizados",
            "codigo": cod,
            "produto": nome or str(prod).strip(),
            "un": "UN",
            "preco": round(preco, 2),
        })
    return itens


# --- ABA 2: Barras ---
# L2: Bloco1 A=sabor B=valor | Bloco2 F=sabor G=valor | Un=KG
def extrair_barras(ws) -> list[dict]:
    itens = []
    for r in _linhas_dados(ws):
        sabor = ws.cell(r, 1).value
        val = ws.cell(r, 2).value
        if sabor and str(sabor).strip() and str(sabor).strip().upper() != "TOTAL":
            preco = _parse_valor(val)
            if preco is None:
                preco = 0.0
            if preco >= 0:
                itens.append({
                    "categoria": "Barras ao Leite",
                    "codigo": "",
                    "produto": str(sabor).strip(),
                    "un": "KG",
                    "preco": round(preco, 2),
                })
    for r in _linhas_dados(ws):
        sabor = ws.cell(r, 6).value
        val = ws.cell(r, 7).value
        if sabor and str(sabor).strip() and str(sabor).strip().upper() != "TOTAL":
            preco = _parse_valor(val)
            if preco is None:
                preco = 0.0
            if preco >= 0:
                itens.append({
                    "categoria": "Barras Branco",
                    "codigo": "",
                    "produto": str(sabor).strip(),
                    "un": "KG",
                    "preco": round(preco, 2),
                })
    return itens


# --- ABA 3: Bombons liquidos ---
# Bloco1: A=tipo B=valor/kg | Bloco2: F=tipo G=valor/und
def extrair_bombons_liquidos(ws) -> list[dict]:
    itens = []
    for r in _linhas_dados(ws):
        tipo = ws.cell(r, 1).value
        val_kg = ws.cell(r, 2).value
        if tipo and str(tipo).strip():
            preco = _parse_valor(val_kg)
            if preco is None:
                preco = 0.0
            if preco >= 0:
                itens.append({
                    "categoria": "Bombons Líquidos",
                    "codigo": "",
                    "produto": f"{str(tipo).strip()} (kg)",
                    "un": "KG",
                    "preco": round(preco, 2),
                })
        tipo2 = ws.cell(r, 6).value
        val_un = ws.cell(r, 7).value
        if tipo2 and str(tipo2).strip():
            preco = _parse_valor(val_un)
            if preco is None:
                preco = 0.0
            if preco >= 0:
                itens.append({
                    "categoria": "Bombons Líquidos",
                    "codigo": "",
                    "produto": f"{str(tipo2).strip()} (un)",
                    "un": "UN",
                    "preco": round(preco, 2),
                })
    return itens


# --- ABA 4: Bombons 12gr ---
# Bloco1: A=tipo B=valor/kg | Bloco2: G=tipo H=valor/und
def extrair_bombons_12gr(ws) -> list[dict]:
    itens = []
    for r in _linhas_dados(ws):
        tipo = ws.cell(r, 1).value
        val_kg = ws.cell(r, 2).value
        if tipo and str(tipo).strip() and "TOTAL" not in str(tipo).upper():
            preco = _parse_valor(val_kg)
            if preco is None:
                preco = 0.0
            if preco >= 0:
                itens.append({
                    "categoria": "Bombons 12gr",
                    "codigo": "",
                    "produto": f"{str(tipo).strip()} (kg)",
                    "un": "KG",
                    "preco": round(preco, 2),
                })
        tipo2 = ws.cell(r, 7).value
        val_un = ws.cell(r, 8).value
        if tipo2 and str(tipo2).strip() and "TOTAL" not in str(tipo2).upper():
            preco = _parse_valor(val_un)
            if preco is None:
                preco = 0.0
            if preco >= 0:
                itens.append({
                    "categoria": "Bombons 12gr",
                    "codigo": "",
                    "produto": f"{str(tipo2).strip()} (un)",
                    "un": "UN",
                    "preco": round(preco, 2),
                })
    return itens


# --- ABA 5: Trufas e trufados ---
# Bloco1: A=sabor B=valor (saco 25) | Bloco2: F=sabor G=valor (und)
def extrair_trufas(ws) -> list[dict]:
    itens = []
    for r in _linhas_dados(ws):
        sabor = ws.cell(r, 1).value
        val = ws.cell(r, 2).value
        if sabor and str(sabor).strip():
            preco = _parse_valor(val)
            if preco is None:
                preco = 0.0
            if preco >= 0:
                itens.append({
                    "categoria": "Trufas",
                    "codigo": "",
                    "produto": f"{str(sabor).strip()} (saco 25)",
                    "un": "SACO",
                    "preco": round(preco, 2),
                })
        sabor2 = ws.cell(r, 6).value
        val2 = ws.cell(r, 7).value
        if sabor2 and str(sabor2).strip():
            preco = _parse_valor(val2)
            if preco is None:
                preco = 0.0
            if preco >= 0:
                itens.append({
                    "categoria": "Trufas",
                    "codigo": "",
                    "produto": f"{str(sabor2).strip()} (un)",
                    "un": "UN",
                    "preco": round(preco, 2),
                })
    return itens


# --- ABA 6: Degustação ---
# A=tipo B=valor (ex: "R$ 99,00 Kg")
def extrair_degustacao(ws) -> list[dict]:
    itens = []
    for r in _linhas_dados(ws):
        tipo = ws.cell(r, 1).value
        val = ws.cell(r, 2).value
        if tipo and str(tipo).strip():
            preco = _parse_valor(val)
            if preco is None:
                preco = 0.0
            if preco >= 0:
                itens.append({
                    "categoria": "Degustação",
                    "codigo": "",
                    "produto": str(tipo).strip(),
                    "un": "KG",
                    "preco": round(preco, 2),
                })
    return itens


# --- ABA 7: Planilha9 ---
# Bloco1 (50% CACAU): A=nome B=preco_kg C=preco_un
# Bloco2 (BOMBOM ESPECIAIS): F=nome G=preco_kg H=preco_un
# Ignorar: 50% CACAU, 70% CACAU, BOMBOM ESPECIAIS (são headers)
def extrair_planilha9(ws) -> list[dict]:
    itens = []
    skip = {"", "50% CACAU", "70% CACAU", "BOMBOM ESPECIAIS"}

    def _bloco_planilha9(nome, val_kg, val_un, categoria: str):
        if not nome or str(nome).strip() in skip:
            return
        pk = _parse_valor(val_kg)
        pu = _parse_valor(val_un)
        if pk is None:
            pk = 0.0
        if pu is None:
            pu = 0.0
        nm = str(nome).strip()
        # Inclui linhas com preço só em kg, só em un, ou ambos; evita duplicar (un) quando igual a (kg)
        if pk > 0:
            itens.append({
                "categoria": categoria,
                "codigo": "",
                "produto": f"{nm} (kg)",
                "un": "KG",
                "preco": round(pk, 2),
            })
        if pu > 0 and abs(pu - pk) > 1e-9:
            itens.append({
                "categoria": categoria,
                "codigo": "",
                "produto": f"{nm} (un)",
                "un": "UN",
                "preco": round(pu, 2),
            })
        elif pu > 0 and pk == 0:
            itens.append({
                "categoria": categoria,
                "codigo": "",
                "produto": f"{nm} (un)",
                "un": "UN",
                "preco": round(pu, 2),
            })
        elif pk == 0 and pu == 0:
            itens.append({
                "categoria": categoria,
                "codigo": "",
                "produto": f"{nm} (kg)",
                "un": "KG",
                "preco": 0.0,
            })

    for r in _linhas_dados(ws):
        _bloco_planilha9(ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value, "50% Cacau")
        _bloco_planilha9(ws.cell(r, 6).value, ws.cell(r, 7).value, ws.cell(r, 8).value, "Bombons Especiais")
    return itens


EXTRATORES = {
    "Personalizados": extrair_personalizados,
    "Barras": extrair_barras,
    "Bombons liquidos": extrair_bombons_liquidos,
    "Bombons 12gr": extrair_bombons_12gr,
    "Trufas e trufados": extrair_trufas,
    "Degustação": extrair_degustacao,
    "Planilha9": extrair_planilha9,
}
# Nomes alternativos comuns no Excel (acentos / maiúsculas)
for _alias, _fn in (
    ("Bombons líquidos", extrair_bombons_liquidos),
    ("Bombons Líquidos", extrair_bombons_liquidos),
    ("BOMBONS LIQUIDOS", extrair_bombons_liquidos),
    ("Trufas", extrair_trufas),
    ("TRUFAS", extrair_trufas),
    ("Degustacao", extrair_degustacao),
    ("DEGUSTAÇÃO", extrair_degustacao),
    ("BARRAS", extrair_barras),
    ("PERSONALIZADOS", extrair_personalizados),
):
    EXTRATORES.setdefault(_alias, _fn)


def _extrator_por_nome_aba(sheet_name: str):
    """
    Resolve extrator mesmo com nome de aba ligeiramente diferente (encoding, maiúsculas, sufixos).
    """
    sn = (sheet_name or "").strip()
    if sn in EXTRATORES:
        return EXTRATORES[sn]
    low = sn.lower()
    low = low.replace("í", "i").replace("ã", "a").replace("ç", "c").replace("õ", "o")
    if "degust" in low:
        return extrair_degustacao
    if "personaliz" in low:
        return extrair_personalizados
    if "planilha" in low and "9" in sn:
        return extrair_planilha9
    if "trufa" in low:
        return extrair_trufas
    if "bombom" in low or "bombons" in low:
        if re.search(r"12\s*gr|12gr", low):
            return extrair_bombons_12gr
        if "liqu" in low or "liqui" in low:
            return extrair_bombons_liquidos
    if "barra" in low:
        return extrair_barras
    return None


def extrair_todos_workbook(wb) -> list[dict]:
    """Extrai produtos a partir de um workbook já aberto (openpyxl)."""
    todos = []
    for sn in wb.sheetnames:
        ext = EXTRATORES.get(sn) or _extrator_por_nome_aba(sn)
        if ext:
            try:
                itens = ext(wb[sn])
                todos.extend(itens)
            except Exception as e:
                print(f"Aviso: erro em aba '{sn}': {e}")
    return todos


def extrair_todos_de_bytes(data: bytes) -> list[dict]:
    """Carrega planilha a partir de bytes (upload na nuvem)."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    try:
        return extrair_todos_workbook(wb)
    finally:
        wb.close()


def extrair_todos(caminho: str | Path | None = None) -> list[dict]:
    """Extrai todos os produtos de todas as abas."""
    import openpyxl

    path = caminho or _caminho_planilha()
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        todos = extrair_todos_workbook(wb)
    finally:
        wb.close()
    return todos
