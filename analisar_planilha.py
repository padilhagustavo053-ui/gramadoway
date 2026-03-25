"""
Analisa cada aba da planilha Gramadoway e gera relatório detalhado.
"""
from pathlib import Path
import openpyxl


def encontrar_planilha():
    desktop = Path.home() / "Desktop"
    for f in desktop.glob("*Gramadoway*.xlsx"):
        return f
    return None


def analisar():
    path = encontrar_planilha()
    if not path:
        print("Planilha não encontrada em Desktop.")
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    print("=" * 70)
    print("ANÁLISE COMPLETA — Planilha preços Gramadoway")
    print("=" * 70)
    print(f"\nArquivo: {path.name}\n")
    for sn in wb.sheetnames:
        ws = wb[sn]
        print("-" * 70)
        print(f"ABA: {sn} | Linhas: {ws.max_row} | Colunas: {ws.max_column}")
        print("-" * 70)
        for r in range(1, min(8, ws.max_row + 1)):
            row_vals = [ws.cell(r, c).value for c in range(1, min(12, ws.max_column + 1))]
            print(f"  L{r}: {row_vals}")
        print()
    wb.close()


if __name__ == "__main__":
    analisar()
