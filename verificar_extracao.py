"""
Verifica se a extração está completa — compara com o que está em cada aba.
"""
from pathlib import Path
import openpyxl
from collections import defaultdict, Counter
from extrair import extrair_todos, EXTRATORES, _caminho_planilha


def main():
    path = _caminho_planilha()
    wb = openpyxl.load_workbook(path, data_only=True)
    
    print("=" * 70)
    print("VERIFICACAO COMPLETA - Extracao vs Planilha Real")
    print("=" * 70)
    print(f"\nArquivo: {path.name}\n")
    
    todos = extrair_todos(path)
    por_cat = defaultdict(list)
    for p in todos:
        por_cat[p["categoria"]].append(p)
    
    cat_map = {
        "Personalizados": "Personalizados",
        "Barras": ("Barras ao Leite", "Barras Branco"),
        "Bombons liquidos": "Bombons Líquidos",
        "Bombons 12gr": "Bombons 12gr",
        "Trufas e trufados": "Trufas",
        "Degustação": "Degustação",
        "Degusta��o": "Degustação",
        "Planilha9": ("50% Cacau", "Bombons Especiais"),
    }
    
    for sn in wb.sheetnames:
        ws = wb[sn]
        if sn not in EXTRATORES and "Degust" not in sn:
            continue
        
        n_linhas = sum(1 for r in range(3, ws.max_row + 1) if ws.cell(r, 1).value or ws.cell(r, 6).value)
        cats = cat_map.get(sn, sn)
        if isinstance(cats, tuple):
            n_extraido = sum(len(por_cat.get(c, [])) for c in cats)
            amostra = []
            for c in cats:
                amostra.extend(por_cat.get(c, [])[:2])
        else:
            n_extraido = len(por_cat.get(cats, []))
            amostra = por_cat.get(cats, [])[:4]
        
        print(f"ABA: {sn}")
        print(f"  Linhas planilha: ~{n_linhas} | Extraidos: {n_extraido}")
        for p in amostra:
            cod = f" [{p['codigo']}]" if p.get("codigo") else ""
            print(f"    -> {p['produto']}{cod} | {p['un']} | R$ {p['preco']}")
        print()
    
    wb.close()
    por_un = Counter(p["un"] for p in todos)
    com_codigo = sum(1 for p in todos if p.get("codigo"))
    print("=" * 70)
    print(f"TOTAL: {len(todos)} produtos | KG={por_un.get('KG',0)} UN={por_un.get('UN',0)} SACO={por_un.get('SACO',0)}")
    print(f"Com codigo: {com_codigo}")
    print("=" * 70)


if __name__ == "__main__":
    main()
