"""
Verifica se a extração está completa — compara com o que está em cada aba.
"""
from pathlib import Path
import openpyxl
from extrair import extrair_todos, EXTRATORES, _caminho_planilha


def main():
    path = _caminho_planilha()
    wb = openpyxl.load_workbook(path, data_only=True)
    
    print("=" * 70)
    print("VERIFICAÇÃO COMPLETA — Extração vs Planilha Real")
    print("=" * 70)
    print(f"\nArquivo: {path.name}\n")
    
    todos = extrair_todos(path)
    
    # Agrupa por categoria (que mapeia às abas)
    from collections import defaultdict
    por_cat = defaultdict(list)
    for p in todos:
        por_cat[p["categoria"]].append(p)
    
    for sn in wb.sheetnames:
        ws = wb[sn]
        ext = EXTRATORES.get(sn)
        if not ext and "Degust" not in sn:
            continue
            
        # Conta linhas com dados na planilha
        n_linhas = 0
        for r in range(3, ws.max_row + 1):
            if ws.cell(r, 1).value or ws.cell(r, 6).value:  # col A ou F
                n_linhas += 1
        
        # Mapeia categoria da aba
        cat_map = {
            "Personalizados": "Personalizados",
            "Barras": "Barras ao Leite",  # + Barras Branco
            "Bombons liquidos": "Bombons Líquidos",
            "Bombons 12gr": "Bombons 12gr",
            "Trufas e trufados": "Trufas",
            "Degustação": "Degustação",
            "Planilha9": "50% Cacau",  # + Bombons Especiais
        }
        cats_aba = cat_map.get(sn, sn)
        if sn == "Barras":
            n_extraido = len(por_cat.get("Barras ao Leite", [])) + len(por_cat.get("Barras Branco", []))
        elif sn == "Planilha9":
            n_extraido = len(por_cat.get("50% Cacau", [])) + len(por_cat.get("Bombons Especiais", []))
        else:
            n_extraido = len(por_cat.get(cats_aba, []))
        
        print(f"ABA: {sn}")
        print(f"  Linhas com dados (planilha): ~{n_linhas}")
        print(f"  Produtos extraídos: {n_extraido}")
        
        # Amostra: primeiros 3 produtos
        amostra = [p for p in todos if p["categoria"] in (cats_aba if isinstance(cats_aba, str) else [cats_aba, "Barras Branco", "Bombons Especiais"])]
        if sn == "Barras":
            amostra = por_cat.get("Barras ao Leite", [])[:2] + por_cat.get("Barras Branco", [])[:2]
        elif sn == "Planilha9":
            amostra = por_cat.get("50% Cacau", [])[:2] + por_cat.get("Bombons Especiais", [])[:2]
        else:
            amostra = por_cat.get(cats_aba, [])[:4]
        
        for p in amostra:
            cod = f" [{p['codigo']}]" if p.get("codigo") else ""
            print(f"    → {p['produto']}{cod} | {p['un']} | R$ {p['preco']}")
        print()
    
    wb.close()
    print("=" * 70)
    print(f"TOTAL EXTRAÍDO: {len(todos)} produtos")
    print("=" * 70)
    
    # Resumo por unidade
    from collections import Counter
    por_un = Counter(p["un"] for p in todos)
    print("\nPor unidade: KG=%d UN=%d SACO=%d" % (por_un.get("KG",0), por_un.get("UN",0), por_un.get("SACO",0)))
    
    # Produtos com código
    com_codigo = sum(1 for p in todos if p.get("codigo"))
    print(f"Produtos com código: {com_codigo} (aba Personalizados)")


if __name__ == "__main__":
    main()
