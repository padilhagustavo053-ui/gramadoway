"""
Gera data/planilha.xlsx com estrutura Gramadoway (abas e colunas que extrair.py espera).
Execute: python scripts/gerar_planilha_embutida.py
Substitua o ficheiro pelos seus preços reais quando quiser; na nuvem o deploy já traz produtos de exemplo.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "data" / "planilha.xlsx"


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    # --- Personalizados (L2 cabeçalho, L3+ dados) ---
    ws = wb.create_sheet("Personalizados")
    ws["A2"], ws["B2"] = "PRODUTO", "VALOR"
    ws["A3"], ws["B3"] = "101 - Ovo P 50g", 12.5
    ws["A4"], ws["B4"] = "102 - Coração 80g", 18.9

    # --- Barras (2 blocos A-B e F-G a partir da linha 3) ---
    ws = wb.create_sheet("Barras")
    ws["A3"], ws["B3"] = "Ao leite tradicional", 89.0
    ws["A4"], ws["B4"] = "Meio amargo", 95.0
    ws["F3"], ws["G3"] = "Branco com nibs", 102.0

    # --- Bombons liquidos ---
    ws = wb.create_sheet("Bombons liquidos")
    ws["A3"], ws["B3"] = "Cereja", 120.0
    ws["F3"], ws["G3"] = "Licor fino", 4.5

    # --- Bombons 12gr ---
    ws = wb.create_sheet("Bombons 12gr")
    ws["A3"], ws["B3"] = "Sortido premium", 140.0
    ws["G3"], ws["H3"] = "Trufa ouro", 3.2

    # --- Trufas e trufados ---
    ws = wb.create_sheet("Trufas e trufados")
    ws["A3"], ws["B3"] = "Trufa tradicional", 85.0
    ws["F3"], ws["G3"] = "Trufa pistache", 5.5

    # --- Degustação ---
    ws = wb.create_sheet("Degustação")
    ws["A3"], ws["B3"] = "Mesa degustação", 150.0

    # --- Planilha9 (50% / bombons especiais) ---
    ws = wb.create_sheet("Planilha9")
    ws["A3"], ws["B3"], ws["C3"] = "Tablete 50%", 88.0, 8.5
    ws["F3"], ws["G3"], ws["H3"] = "Bombom especial", 160.0, 6.0

    wb.save(OUT)
    wb.close()
    print(f"OK: {OUT}")


if __name__ == "__main__":
    main()
