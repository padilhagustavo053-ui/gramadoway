"""
Teste automático: planilha sintética com muitas linhas → extrair.py deve trazer todas.
Execute na pasta do projeto: python scripts/test_extracao_bulk.py
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402

from extrair import extrair_todos  # noqa: E402


def _montar_workbook(n_personalizados: int = 180) -> Path:
    wb = openpyxl.Workbook()
    # Personalizados
    ws = wb.active
    ws.title = "Personalizados"
    ws.cell(1, 1, "TITULO")
    ws.cell(2, 1, "PRODUTO")
    ws.cell(2, 2, "VALOR")
    for i in range(n_personalizados):
        r = 3 + i
        ws.cell(r, 1, f"{500 + i}- Produto teste {i}")
        ws.cell(r, 2, 1.5 + (i % 10) * 0.25)
    # Barras mínimo (evitar aba vazia estranha)
    wb.create_sheet("Barras")
    wb["Barras"].cell(2, 1, "Sabor")
    wb["Barras"].cell(2, 2, "Valor")
    wb["Barras"].cell(3, 1, "Teste KG")
    wb["Barras"].cell(3, 2, 88.0)

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    p = Path(tmp.name)
    tmp.close()
    wb.save(p)
    wb.close()
    return p


def main() -> None:
    n = 180
    path = _montar_workbook(n)
    try:
        itens = extrair_todos(path)
        pers = [x for x in itens if x["categoria"] == "Personalizados"]
        assert len(pers) == n, f"Esperado {n} Personalizados, obteve {len(pers)}"
        print(f"OK — {len(pers)} personalizados extraídos (teste bulk). Total linhas produto: {len(itens)}")
    finally:
        path.unlink(missing_ok=True)


if __name__ == "__main__":
    main()
